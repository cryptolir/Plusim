# -*- coding: utf-8 -*-
"""Build the month-sheet workbook (reference/layout-spec.md).

Sheets, in order: Main (taxonomy list) · one RTL sheet per month in the example
layout (section headers with subtotal formulas, leaf rows with amounts spread
across columns + a SUM formula, grand-total row, manual-fill income block) ·
ניתוח תוצאות (cross-month analysis) · התפלגות ההוצאות (section averages + pie
chart) · טופס עזר למיפוי (static advisor intake form) · חישוב יעדים (savings
goal calculator) · un_categorized · פירוט תנועות (full audit ledger).

Amounts arrive as agorot ints and are written as shekels with the ₪ number
format; totals are live formulas so the sheet stays hand-editable. Every
cross-sheet reference is generated from a per-sheet geometry map returned by
the sheet builders — never a hardcoded address, because the month-sheet total
column is dynamic (docs/plans/report-workbook-subreports.md).
"""
from __future__ import annotations

import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

CURR = "₪ #,##0.00"
HDR_FILL = PatternFill("solid", fgColor="4472C4")
GRAND_FILL = PatternFill("solid", fgColor="FFF2CC")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
WHITE_BOLD = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)

HEB_MONTHS = [
    "ינואר", "פברואר", "מרץ", "אפריל", "מאי", "יוני",
    "יולי", "אוגוסט", "ספטמבר", "אוקטובר", "נובמבר", "דצמבר",
]

# Manual-fill income rows appended to every month sheet (plan A2).
INCOME_LABELS = ["משכורת 1", "הכנסות נוספות", "משכורת 2"]

# Business/tax leaves excluded from שונות in the distribution pie (Decision 3).
BUSINESS_EXCLUDE_LEAVES = ("עסק", "מעמ", "מס הכנסה", "ביטוח לאומי")

ANALYSIS_SHEET = "ניתוח תוצאות"
DISTRIBUTION_SHEET = "התפלגות ההוצאות"
MAPPING_FORM_SHEET = "טופס עזר למיפוי"
GOALS_SHEET = "חישוב יעדים"

_CARD_NAMES = ["ויזה כאל ", "מקס לאומי קארד", "ישראכרט", "דיינרס", "AMX"]


def _form_cells() -> list[tuple[int, int, str]]:
    """טופס עזר למיפוי content, transcribed verbatim from the demo workbook
    (plan appendix — including the template's trailing spaces)."""
    cells: list[tuple[int, int, str]] = [
        (1, 1, "כמות קופת חולים "),
        (2, 1, "שם קופת חולים "),
        (4, 1, "ביטוחים חיים בריאות "),
        (15, 1, "ביטוחים אלמנטרי- רכב דירה "),
        (26, 1, "הוראות קבע שיורדות מהבנק "),
        (32, 1, "ריבוי משיכות  מכספומט פרטי "), (32, 2, "לא"),
        (34, 1, "חיובי תקשורת גבוהים "), (34, 2, "לא"), (34, 3, "חברות תקשורת "),
        (36, 1, "קצבת ילדים"), (36, 2, "לא"), (36, 3, "סכום קיצבה משולם"),
        (36, 5, "סכום קיצבה מקורי "), (36, 7, "חיסכון לכל ילד "),
        (38, 1, "חיובי עמלות גבוהים "), (38, 2, "לא"), (38, 6, "כמות כרטיסי אשראי "),
        (40, 1, "הלוואות בנקאיות "), (40, 2, "סכום ההלוואה "), (40, 3, "תשלום חודשי "),
        (40, 4, "אחוז ריבית "), (40, 6, "כרטיסי אשראי "), (40, 7, "כמות "),
        (50, 1, "הלוואות חוץ בנקאיות "), (50, 2, "סכום ההלוואה "),
        (50, 3, "תשלום חודשי "), (50, 4, "אחוז ריבית "),
    ]
    for col in range(2, 9):
        cells.append((4, col, "סכום "))
        cells.append((15, col, "סכום "))
    for i, name in enumerate(
        ["הראל", "מגדל", "כלל", "פניקס", "מנורה ", "הכשרת הביטוח", "AIG", "ביטוח ישיר"]
    ):
        cells.append((5 + i, 1, name))
    for i, name in enumerate(
        ["הראל", "מגדל", "כלל", "פניקס", "מנורה ", "הכשרת הביטוח", "איילון", "AIG", "ביטוח ישיר"]
    ):
        cells.append((16 + i, 1, name))
    for i, name in enumerate(
        ["דיסקונט", "פועלים", "מסד", "יהב", "לאומי", "בנק ירושלים", "מזרחי", "אוצר החייל "]
    ):
        cells.append((41 + i, 1, name))
    for i, name in enumerate(_CARD_NAMES):
        cells.append((41 + i, 6, name))
    for i, name in enumerate(
        _CARD_NAMES
        + [
            "ליסינג בחברת ___________",
            "מימון ישיר ",
            "פמה ",
            "הלוואה חברתית ב__________",
            'גמ"ח ____________________',
        ]
    ):
        cells.append((51 + i, 1, name))
    return cells


MAPPING_FORM_CELLS = _form_cells()


def month_title(month: str) -> str:
    y, m = month.split("-")
    return f"{HEB_MONTHS[int(m) - 1]} {y}"


def _shekels(agorot: int) -> float:
    return agorot / 100.0


def build_workbook(taxonomy: list[dict], txns: list[dict], out_path: str) -> None:
    """taxonomy: [{"section": str, "leaves": [str]}]; txns: normalized dicts
    with month/date/merchant/amountAgorot/category/uncategorized/sourceLabel/note."""
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Main"
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "Main - רשימת קטגוריות"
    ws["A1"].font = TITLE_FONT
    r = 3
    for sec in taxonomy:
        c = ws.cell(r, 1, sec["section"])
        c.font = WHITE_BOLD
        c.fill = HDR_FILL
        r += 1
        for leaf in sec["leaves"]:
            ws.cell(r, 1, leaf)
            r += 1
    ws.column_dimensions["A"].width = 30

    months = sorted({t["month"] for t in txns})
    geoms = []
    for month in months:
        geoms.append(_month_sheet(wb, taxonomy, [t for t in txns if t["month"] == month], month))

    if geoms:
        analysis = _analysis_sheet(wb, taxonomy, geoms)
        _distribution_sheet(wb, taxonomy, analysis)
    _mapping_form_sheet(wb)
    _goals_sheet(wb)

    _uncat_sheet(wb, [t for t in txns if t["uncategorized"]])
    _ledger_sheet(wb, txns)
    wb.save(out_path)


def _month_sheet(wb, taxonomy, txns, month: str) -> dict:
    """Build one month sheet; return its geometry map (plan A1)."""
    ws = wb.create_sheet(month_title(month))
    ws.sheet_view.rightToLeft = True
    ws.cell(1, 1, "חודש " + month_title(month)).font = TITLE_FONT

    amounts: dict[str, list[int]] = {}
    for t in txns:
        if not t["uncategorized"] and t.get("category"):
            amounts.setdefault(t["category"], []).append(t["amountAgorot"])

    maxk = max([len(v) for v in amounts.values()] + [9])
    tcol = maxk + 2
    tl = get_column_letter(tcol)

    row = 3
    header_rows: list[int] = []
    section_rows: dict[str, int] = {}
    leaf_rows: dict[str, int] = {}
    cur_leaf_rows: list[int] = []
    pending_header: int | None = None

    def close_section():
        if pending_header is not None:
            cell = ws.cell(pending_header, tcol)
            cell.value = ("=" + "+".join(f"{tl}{r}" for r in cur_leaf_rows)) if cur_leaf_rows else 0
            cell.font = WHITE_BOLD
            cell.fill = HDR_FILL
            cell.number_format = CURR

    for sec in taxonomy:
        close_section()
        cur_leaf_rows = []
        pending_header = row
        h = ws.cell(row, 1, sec["section"])
        h.font = WHITE_BOLD
        h.fill = HDR_FILL
        header_rows.append(row)
        section_rows[sec["section"]] = row
        row += 1
        for leaf in sec["leaves"]:
            ws.cell(row, 1, leaf)
            for i, agorot in enumerate(amounts.get(leaf, [])):
                c = ws.cell(row, 2 + i, _shekels(agorot))
                c.number_format = CURR
            tc = ws.cell(row, tcol, f"=SUM(B{row}:{get_column_letter(tcol - 1)}{row})")
            tc.number_format = CURR
            tc.font = BOLD
            cur_leaf_rows.append(row)
            leaf_rows[leaf] = row
            row += 1
    close_section()

    grand = row + 1
    ws.cell(grand, 1, 'סה"כ החודש').font = BOLD
    gc = ws.cell(grand, tcol, "=" + "+".join(f"{tl}{r}" for r in header_rows))
    gc.number_format = CURR
    gc.font = Font(bold=True, size=12)
    gc.fill = GRAND_FILL

    # Manual-fill income block (plan A2; template rows 64-71 normalized: income
    # totals live in the sheet's single total column, plain SUMs only).
    inc_hdr = grand + 2
    ws.cell(inc_hdr, 1, "הכנסות :").font = BOLD
    income_rows: dict[str, int] = {}
    r = inc_hdr + 1
    for label in INCOME_LABELS:
        ws.cell(r, 1, label)
        c = ws.cell(r, tcol, f"=SUM(B{r}:{get_column_letter(tcol - 1)}{r})")
        c.number_format = CURR
        income_rows[label] = r
        r += 1
    income_total = r
    ws.cell(income_total, 1, 'סה"כ').font = BOLD
    c = ws.cell(income_total, tcol, f"=SUM({tl}{inc_hdr + 1}:{tl}{income_total - 1})")
    c.number_format = CURR
    c.font = BOLD

    net_first = income_total + 2
    ws.cell(net_first, 1, "הכנסות").font = BOLD
    c = ws.cell(net_first, 2, f"={tl}{income_total}")
    c.number_format = CURR
    ws.cell(net_first + 1, 1, "הוצאות").font = BOLD
    c = ws.cell(net_first + 1, 2, f"=-{tl}{grand}")
    c.number_format = CURR
    ws.cell(net_first + 2, 1, "זכות /חובה").font = BOLD
    c = ws.cell(net_first + 2, 2, f"=SUM(B{net_first}:B{net_first + 1})")
    c.number_format = CURR

    ws.column_dimensions["A"].width = 26
    for ci in range(2, tcol + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 12

    return {
        "sheet": ws.title,
        "tcol": tcol,
        "section_rows": section_rows,
        "leaf_rows": leaf_rows,
        "grand_row": grand,
        "income_rows": income_rows,
        "income_total_row": income_total,
    }


def _analysis_sheet(wb, taxonomy, geoms: list[dict]) -> dict:
    """ניתוח תוצאות — cross-month comparison, built purely from the month
    geometry maps; returns its own geometry (plan B1)."""
    ws = wb.create_sheet(ANALYSIS_SHEET)
    ws.sheet_view.rightToLeft = True
    ws["A1"] = ANALYSIS_SHEET
    ws["A1"].font = TITLE_FONT

    n = len(geoms)
    avg_col = 2 + n
    ac = get_column_letter(avg_col)
    last_month_col = get_column_letter(avg_col - 1)

    def ref(g: dict, row: int) -> str:
        return f"='{g['sheet']}'!{get_column_letter(g['tcol'])}{row}"

    def avg_formula(r: int) -> str:
        # Row-qualified — a whole-column AVERAGE(B:E) would mix rows (Codex round 4).
        return f"=AVERAGE(B{r}:{last_month_col}{r})"

    hdr = 3
    for i, g in enumerate(geoms):
        ws.cell(hdr, 2 + i, g["sheet"]).font = BOLD
    ws.cell(hdr, avg_col, f"ממוצע {n} חודשים").font = BOLD

    section_rows: dict[str, int] = {}
    leaf_rows: dict[str, int] = {}
    r = hdr + 1
    for sec in taxonomy:
        section_rows[sec["section"]] = r
        for col in range(1, avg_col + 1):
            ws.cell(r, col).fill = YELLOW_FILL
        ws.cell(r, 1, sec["section"]).font = BOLD
        for i, g in enumerate(geoms):
            c = ws.cell(r, 2 + i, ref(g, g["section_rows"][sec["section"]]))
            c.number_format = CURR
        c = ws.cell(r, avg_col, avg_formula(r))
        c.number_format = CURR
        r += 1
        for leaf in sec["leaves"]:
            leaf_rows[leaf] = r
            ws.cell(r, 1, leaf).font = BOLD
            for i, g in enumerate(geoms):
                c = ws.cell(r, 2 + i, ref(g, g["leaf_rows"][leaf]))
                c.number_format = CURR
            c = ws.cell(r, avg_col, avg_formula(r))
            c.number_format = CURR
            r += 1

    expense_total_row = r
    for col in range(1, avg_col + 1):
        ws.cell(expense_total_row, col).fill = YELLOW_FILL
    ws.cell(expense_total_row, 1, 'סה"כ').font = BOLD
    for i, g in enumerate(geoms):
        c = ws.cell(
            expense_total_row,
            2 + i,
            f"=-'{g['sheet']}'!{get_column_letter(g['tcol'])}{g['grand_row']}",
        )
        c.number_format = CURR
    c = ws.cell(expense_total_row, avg_col, avg_formula(expense_total_row))
    c.number_format = CURR

    # Income comparison (plan B1; both blocks share one month-column order).
    inc_hdr = expense_total_row + 2
    ws.cell(inc_hdr, 1, "כמה כסף הבאנו הביתה").font = BOLD
    for i, g in enumerate(geoms):
        ws.cell(inc_hdr, 2 + i, g["sheet"]).font = BOLD
    ws.cell(inc_hdr, avg_col, f"ממוצע {n} חודשים").font = BOLD
    r = inc_hdr + 1
    for label in INCOME_LABELS:
        ws.cell(r, 1, label)
        for i, g in enumerate(geoms):
            c = ws.cell(r, 2 + i, ref(g, g["income_rows"][label]))
            c.number_format = CURR
        c = ws.cell(r, avg_col, avg_formula(r))
        c.number_format = CURR
        r += 1
    income_total_row = r
    ws.cell(income_total_row, 1, 'סה"כ').font = BOLD
    for col in range(2, avg_col + 1):
        cl = get_column_letter(col)
        c = ws.cell(income_total_row, col, f"=SUM({cl}{inc_hdr + 1}:{cl}{income_total_row - 1})")
        c.number_format = CURR

    # מצבנו — averages of income vs expenses.
    st_hdr = income_total_row + 2
    ws.cell(st_hdr, 1, "מצבנו").font = BOLD
    ws.cell(st_hdr + 1, 1, "הכנסות")
    c = ws.cell(st_hdr + 1, 2, f"={ac}{income_total_row}")
    c.number_format = CURR
    ws.cell(st_hdr + 2, 1, "הוצאות")
    c = ws.cell(st_hdr + 2, 2, f"={ac}{expense_total_row}")
    c.number_format = CURR
    ws.cell(st_hdr + 3, 1, "יתרה")
    c = ws.cell(st_hdr + 3, 2, f"=SUM(B{st_hdr + 1}:B{st_hdr + 2})")
    c.number_format = CURR

    ws.column_dimensions["A"].width = 26
    for ci in range(2, avg_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    return {
        "sheet": ws.title,
        "avg_col": ac,
        "section_rows": section_rows,
        "leaf_rows": leaf_rows,
        "expense_total_row": expense_total_row,
        "income_total_row": income_total_row,
    }


def _distribution_sheet(wb, taxonomy, analysis: dict) -> None:
    """התפלגות ההוצאות — one row per section referencing the analysis sheet's
    average column via its geometry map, plus a native pie chart (plan B2)."""
    ws = wb.create_sheet(DISTRIBUTION_SHEET)
    ws.sheet_view.rightToLeft = True
    a = analysis["avg_col"]
    sheet = analysis["sheet"]

    first = 2
    r = first
    for sec in taxonomy:
        srow = analysis["section_rows"][sec["section"]]
        ws.cell(r, 1, f"='{sheet}'!A{srow}")
        if sec["section"] == "שונות":
            parts = [f"'{sheet}'!{a}{srow}"]
            for leaf in BUSINESS_EXCLUDE_LEAVES:
                if leaf in sec["leaves"] and leaf in analysis["leaf_rows"]:
                    parts.append(f"'{sheet}'!{a}{analysis['leaf_rows'][leaf]}")
            formula = "=" + "-".join(parts)
        else:
            formula = f"='{sheet}'!{a}{srow}"
        c = ws.cell(r, 2, formula)
        c.number_format = CURR
        r += 1
    last = r - 1

    pie = PieChart()
    data = Reference(ws, min_col=2, min_row=first, max_row=last)
    labels = Reference(ws, min_col=1, min_row=first, max_row=last)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showCatName = True
    pie.dataLabels.showPercent = True
    pie.legend = None  # template chart has no legend — labels carry the names
    ws.add_chart(pie, "G6")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14


def _mapping_form_sheet(wb) -> None:
    """טופס עזר למיפוי — static advisor intake form (plan B3, appendix)."""
    ws = wb.create_sheet(MAPPING_FORM_SHEET)
    ws.sheet_view.rightToLeft = True
    for row, col, text in MAPPING_FORM_CELLS:
        ws.cell(row, col, text)
    ws.column_dimensions["A"].width = 34
    for ci in range(2, 9):
        ws.column_dimensions[get_column_letter(ci)].width = 14


def _goals_sheet(wb) -> None:
    """חישוב יעדים — blank fill-in savings calculator. Every downstream formula
    is anchored on the row's own age input so unused rows stay blank and add
    nothing to the total (plan B4, Codex round 3)."""
    ws = wb.create_sheet(GOALS_SHEET)
    ws.sheet_view.rightToLeft = True
    ws.cell(1, 3, "סכום היעד ")
    ws.cell(1, 4).number_format = CURR
    ws.cell(1, 5, "גיל היעד ")
    for r in range(3, 9):
        ws.cell(r, 1, "שם ")
        ws.cell(r, 3, "גיל ")
        ws.cell(r, 5, "תקופה בשנים ")
        ws.cell(r, 6, f'=IF(D{r}="","",$F$1-D{r})')
        ws.cell(r, 7, "חודשי חיסכון")
        ws.cell(r, 8, f'=IF(D{r}="","",F{r}*12)')
        ws.cell(r, 9, 'סה"כ חיסכון חודשי ')
        c = ws.cell(r, 10, f'=IF(D{r}="","",IF(H{r}>0,$D$1/H{r},""))')
        c.number_format = CURR
    ws.cell(10, 9, 'סה"כ').font = BOLD
    c = ws.cell(10, 10, "=SUM(J3:J8)")
    c.number_format = CURR
    c.font = BOLD
    for ci, w in ((1, 8), (2, 12), (3, 8), (4, 8), (5, 14), (6, 8), (7, 14), (8, 8), (9, 18), (10, 14)):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _uncat_sheet(wb, txns) -> None:
    ws = wb.create_sheet("un_categorized")
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "עסקאות ללא סיווג"
    ws["A1"].font = TITLE_FONT
    for j, h in enumerate(["חודש", "תאריך", "שם בית עסק", "סכום", "מקור", "הערה"], 1):
        c = ws.cell(3, j, h)
        c.font = WHITE_BOLD
        c.fill = HDR_FILL
    r = 4
    for t in txns:
        ws.cell(r, 1, month_title(t["month"]))
        ws.cell(r, 2, t["date"])
        ws.cell(r, 3, t["merchant"])
        a = ws.cell(r, 4, _shekels(t["amountAgorot"]))
        a.number_format = CURR
        ws.cell(r, 5, t["sourceLabel"])
        ws.cell(r, 6, t.get("note") or "")
        r += 1
    for j, w in enumerate([12, 11, 28, 12, 20, 30], 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def _ledger_sheet(wb, txns) -> None:
    ws = wb.create_sheet("פירוט תנועות")
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "פירוט כל התנועות והסיווג"
    ws["A1"].font = TITLE_FONT
    for j, h in enumerate(["חודש", "תאריך", "שם בית עסק", "סכום", "קטגוריה", "מקור", "הערה"], 1):
        c = ws.cell(3, j, h)
        c.font = WHITE_BOLD
        c.fill = HDR_FILL
    r = 4
    for t in sorted(txns, key=lambda x: (x["month"], x["date"])):
        ws.cell(r, 1, month_title(t["month"]))
        ws.cell(r, 2, t["date"])
        ws.cell(r, 3, t["merchant"])
        a = ws.cell(r, 4, _shekels(t["amountAgorot"]))
        a.number_format = CURR
        ws.cell(r, 5, "ללא סיווג" if t["uncategorized"] else t.get("category") or "")
        ws.cell(r, 6, t["sourceLabel"])
        ws.cell(r, 7, t.get("note") or "")
        r += 1
    for j, w in enumerate([12, 11, 28, 12, 22, 20, 30], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
