# -*- coding: utf-8 -*-
"""Named tests for build_report_xlsx.py (docs/plans/report-workbook-subreports.md,
Verification section). Runs on bare python3 with the vendored deps:

    python3 -m unittest agent/skills/plusim-reports/scripts/test_build_report_xlsx.py

Synthetic transactions only — no PII. The verified-baseline re-run with real
statements (docs/REPORTS_PIPELINE.md) remains a separate operator step.
"""
from __future__ import annotations

import base64
import os
import sys
import tempfile
import unittest

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(os.path.dirname(_HERE), "vendor"))
sys.path.insert(0, _HERE)

import openpyxl  # noqa: E402
from openpyxl.chart import PieChart  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from build_report_xlsx import (  # noqa: E402
    ANALYSIS_SHEET,
    BUSINESS_EXCLUDE_LEAVES,
    DISTRIBUTION_SHEET,
    GOALS_SHEET,
    INCOME_LABELS,
    MAPPING_FORM_CELLS,
    MAPPING_FORM_SHEET,
    build_workbook,
    month_title,
)

TAXONOMY = [
    {"section": "בית", "leaves": ["חשמל", "מזון ומכולת"]},
    {"section": "תחבורה", "leaves": ["דלק"]},
    {"section": "שונות", "leaves": ["עסק", "מעמ", "מס הכנסה", "ביטוח לאומי", "מזומן ללא מעקב"]},
]


def txn(month, day, merchant, agorot, category=None, uncat=False):
    return {
        "month": month,
        "date": f"{month}-{day:02d}",
        "merchant": merchant,
        "amountAgorot": agorot,
        "category": category,
        "uncategorized": uncat,
        "sourceLabel": "test-billed",
        "note": "",
    }


def sample_txns(months=("2026-05", "2026-06")):
    out = []
    for m in months:
        out.append(txn(m, 1, "חברת חשמל", 10000, "חשמל"))
        out.append(txn(m, 2, "שופרסל", 20000, "מזון ומכולת"))
        out.append(txn(m, 3, "פז", 15000, "דלק"))
        out.append(txn(m, 4, "מסתורי", 500, None, uncat=True))
    return out


def build(taxonomy, txns):
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    build_workbook(taxonomy, txns, path)
    wb = openpyxl.load_workbook(path)
    return wb, path


def col_a_row(ws, label, after=0):
    """First row after `after` whose column-A value equals label."""
    for r in range(after + 1, ws.max_row + 1):
        if ws.cell(r, 1).value == label:
            return r
    raise AssertionError(f"label {label!r} not found in {ws.title} after row {after}")


def month_geometry(ws, taxonomy):
    """Independently recompute the month sheet's geometry from its contents."""
    tcol = None
    first_leaf = taxonomy[0]["leaves"][0]
    lr = col_a_row(ws, first_leaf)
    for c in range(ws.max_column, 1, -1):
        v = ws.cell(lr, c).value
        if isinstance(v, str) and v.startswith("=SUM("):
            tcol = c
            break
    assert tcol, "total column not found"
    rows = {}
    for sec in taxonomy:
        rows[sec["section"]] = col_a_row(ws, sec["section"])
        for leaf in sec["leaves"]:
            rows[leaf] = col_a_row(ws, leaf)
    grand = col_a_row(ws, 'סה"כ החודש')
    income = {lbl: col_a_row(ws, lbl, after=grand) for lbl in INCOME_LABELS}
    income_total = col_a_row(ws, 'סה"כ', after=max(income.values()))
    return tcol, rows, grand, income, income_total


class BuildWorkbookTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.wb, cls.path = build(TAXONOMY, sample_txns())
        cls.months = [month_title("2026-05"), month_title("2026-06")]

    @classmethod
    def tearDownClass(cls):
        os.unlink(cls.path)

    # -- plan: test_sheet_order_and_rtl ------------------------------------
    def test_sheet_order_and_rtl(self):
        self.assertEqual(
            self.wb.sheetnames,
            ["Main", *self.months, ANALYSIS_SHEET, DISTRIBUTION_SHEET,
             MAPPING_FORM_SHEET, GOALS_SHEET, "un_categorized", "פירוט תנועות"],
        )
        for name in [ANALYSIS_SHEET, DISTRIBUTION_SHEET, MAPPING_FORM_SHEET, GOALS_SHEET]:
            self.assertTrue(self.wb[name].sheet_view.rightToLeft, name)

    # -- plan: test_existing_sheets_unchanged_regression -------------------
    def test_existing_sheets_unchanged_regression(self):
        main = self.wb["Main"]
        expected = []
        for sec in TAXONOMY:
            expected.append(sec["section"])
            expected.extend(sec["leaves"])
        got = [main.cell(r, 1).value for r in range(3, 3 + len(expected))]
        self.assertEqual(got, expected)

        ws = self.wb[self.months[0]]
        self.assertEqual(ws.cell(1, 1).value, "חודש " + self.months[0])
        tcol, rows, grand, _, _ = month_geometry(ws, TAXONOMY)
        self.assertEqual(tcol, 11)  # max 9 amount slots + 2, unchanged default
        tl = get_column_letter(tcol)
        # leaf rows keep the =SUM(B<r>:<tcol-1><r>) shape
        for sec in TAXONOMY:
            for leaf in sec["leaves"]:
                r = rows[leaf]
                self.assertEqual(
                    ws.cell(r, tcol).value,
                    f"=SUM(B{r}:{get_column_letter(tcol - 1)}{r})",
                )
            # section subtotal sums its leaf total cells
            sr = rows[sec["section"]]
            self.assertEqual(
                ws.cell(sr, tcol).value,
                "=" + "+".join(f"{tl}{rows[leaf]}" for leaf in sec["leaves"]),
            )
        self.assertEqual(
            ws.cell(grand, tcol).value,
            "=" + "+".join(f"{tl}{rows[s['section']]}" for s in TAXONOMY),
        )
        # amounts landed as shekels
        self.assertEqual(ws.cell(rows["חשמל"], 2).value, 100.0)

        uncat = self.wb["un_categorized"]
        self.assertEqual(uncat.cell(4, 3).value, "מסתורי")
        ledger = self.wb["פירוט תנועות"]
        self.assertEqual(
            [ledger.cell(3, j).value for j in range(1, 8)],
            ["חודש", "תאריך", "שם בית עסק", "סכום", "קטגוריה", "מקור", "הערה"],
        )

    # -- plan: test_month_income_block_rows_and_formulas -------------------
    def test_month_income_block_rows_and_formulas(self):
        ws = self.wb[self.months[0]]
        tcol, _, grand, income, income_total = month_geometry(ws, TAXONOMY)
        tl = get_column_letter(tcol)
        self.assertEqual(col_a_row(ws, "הכנסות :"), grand + 2)
        for lbl in INCOME_LABELS:
            r = income[lbl]
            self.assertEqual(
                ws.cell(r, tcol).value,
                f"=SUM(B{r}:{get_column_letter(tcol - 1)}{r})",
            )
        self.assertEqual(
            ws.cell(income_total, tcol).value,
            f"=SUM({tl}{min(income.values())}:{tl}{max(income.values())})",
        )
        net_first = col_a_row(ws, "הכנסות", after=income_total)
        self.assertEqual(ws.cell(net_first, 2).value, f"={tl}{income_total}")
        self.assertEqual(ws.cell(net_first + 1, 2).value, f"=-{tl}{grand}")
        self.assertEqual(ws.cell(net_first + 2, 2).value, f"=SUM(B{net_first}:B{net_first + 1})")

    # -- plan: test_analysis_cells_reference_month_geometry ----------------
    def test_analysis_cells_reference_month_geometry(self):
        # one month where a leaf has >9 txns, so its tcol shifts off column K
        txns = sample_txns(months=("2026-05",))
        for d in range(1, 13):
            txns.append(txn("2026-06", d, "שופרסל", 1000, "מזון ומכולת"))
        wb, path = build(TAXONOMY, txns)
        try:
            names = [month_title("2026-05"), month_title("2026-06")]
            geo = {n: month_geometry(wb[n], TAXONOMY) for n in names}
            self.assertEqual(geo[names[0]][0], 11)   # 9 slots default
            self.assertEqual(geo[names[1]][0], 14)   # 12 txns + 2
            an = wb[ANALYSIS_SHEET]
            for i, n in enumerate(names):
                tcol, rows, grand, income, _ = geo[n]
                tl = get_column_letter(tcol)
                for sec in TAXONOMY:
                    r = col_a_row(an, sec["section"], after=3)
                    self.assertEqual(an.cell(r, 2 + i).value, f"='{n}'!{tl}{rows[sec['section']]}")
                    for leaf in sec["leaves"]:
                        lr = col_a_row(an, leaf, after=3)
                        self.assertEqual(an.cell(lr, 2 + i).value, f"='{n}'!{tl}{rows[leaf]}")
                total = col_a_row(an, 'סה"כ', after=3)
                self.assertEqual(an.cell(total, 2 + i).value, f"=-'{n}'!{tl}{grand}")
        finally:
            os.unlink(path)

    # -- plan: test_analysis_average_and_single_month ----------------------
    def test_analysis_average_and_single_month(self):
        an = self.wb[ANALYSIS_SHEET]
        # every average is row-qualified over exactly the month columns (B..C here)
        for r in range(4, an.max_row + 1):
            v = an.cell(r, 4).value  # avg col for 2 months = D
            if isinstance(v, str) and v.startswith("=AVERAGE"):
                self.assertEqual(v, f"=AVERAGE(B{r}:C{r})")
        # single-month job builds a valid sheet with avg col C
        wb, path = build(TAXONOMY, sample_txns(months=("2026-05",)))
        try:
            an1 = wb[ANALYSIS_SHEET]
            r = col_a_row(an1, TAXONOMY[0]["section"], after=3)
            self.assertEqual(an1.cell(r, 3).value, f"=AVERAGE(B{r}:B{r})")
        finally:
            os.unlink(path)

    # -- plan: test_analysis_income_and_status_blocks ----------------------
    def test_analysis_income_and_status_blocks(self):
        an = self.wb[ANALYSIS_SHEET]
        ws = self.wb[self.months[0]]
        _, _, _, income, _ = month_geometry(ws, TAXONOMY)
        inc_hdr = col_a_row(an, "כמה כסף הבאנו הביתה")
        for k, lbl in enumerate(INCOME_LABELS):
            r = inc_hdr + 1 + k
            self.assertEqual(an.cell(r, 1).value, lbl)
            self.assertEqual(an.cell(r, 2).value, f"='{self.months[0]}'!K{income[lbl]}")
        inc_total = inc_hdr + 1 + len(INCOME_LABELS)
        self.assertEqual(an.cell(inc_total, 2).value, f"=SUM(B{inc_hdr + 1}:B{inc_total - 1})")
        st = col_a_row(an, "מצבנו")
        exp_total = col_a_row(an, 'סה"כ', after=3)
        self.assertEqual(an.cell(st + 1, 2).value, f"=D{inc_total}")
        self.assertEqual(an.cell(st + 2, 2).value, f"=D{exp_total}")
        self.assertEqual(an.cell(st + 3, 2).value, f"=SUM(B{st + 1}:B{st + 2})")

    # -- plan: test_distribution_rows_follow_taxonomy ----------------------
    def test_distribution_rows_follow_taxonomy(self):
        # a variant taxonomy without עסק (and without a שונות section at all)
        taxonomy = [
            {"section": "בית", "leaves": ["חשמל"]},
            {"section": "תחבורה", "leaves": ["דלק"]},
        ]
        wb, path = build(taxonomy, [txn("2026-05", 1, "חברת חשמל", 1000, "חשמל"),
                                    txn("2026-05", 2, "פז", 2000, "דלק")])
        try:
            dist = wb[DISTRIBUTION_SHEET]
            an = wb[ANALYSIS_SHEET]
            rows = [r for r in range(1, dist.max_row + 1) if dist.cell(r, 1).value]
            self.assertEqual(len(rows), len(taxonomy))
            for r, sec in zip(rows, taxonomy):
                srow = col_a_row(an, sec["section"], after=3)
                self.assertEqual(dist.cell(r, 1).value, f"='{ANALYSIS_SHEET}'!A{srow}")
                self.assertEqual(dist.cell(r, 2).value, f"='{ANALYSIS_SHEET}'!C{srow}")
        finally:
            os.unlink(path)

    # -- plan: test_distribution_business_exclusion_formula ----------------
    def test_distribution_business_exclusion_formula(self):
        dist = self.wb[DISTRIBUTION_SHEET]
        an = self.wb[ANALYSIS_SHEET]
        srow = col_a_row(an, "שונות", after=3)
        r = 2 + len(TAXONOMY) - 1  # שונות is the last section
        expected = f"='{ANALYSIS_SHEET}'!D{srow}"
        for leaf in BUSINESS_EXCLUDE_LEAVES:
            lr = col_a_row(an, leaf, after=3)
            expected += f"-'{ANALYSIS_SHEET}'!D{lr}"
        self.assertEqual(dist.cell(r, 2).value, expected)

    # -- plan: test_distribution_follows_dynamic_average_column ------------
    def test_distribution_follows_dynamic_average_column(self):
        for months, avg in ((("2026-01",), "C"),
                            (("2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06"), "H")):
            wb, path = build(TAXONOMY, sample_txns(months=months))
            try:
                dist = wb[DISTRIBUTION_SHEET]
                for r in range(2, 2 + len(TAXONOMY)):
                    v = dist.cell(r, 2).value
                    self.assertIn(f"'!{avg}", v, f"months={len(months)}: {v}")
                chart = dist._charts[0]
                self.assertEqual(len(chart.series), 1)
            finally:
                os.unlink(path)

    # -- plan: test_distribution_pie_chart ---------------------------------
    def test_distribution_pie_chart(self):
        dist = self.wb[DISTRIBUTION_SHEET]
        self.assertEqual(len(dist._charts), 1)
        chart = dist._charts[0]
        self.assertIsInstance(chart, PieChart)
        self.assertTrue(chart.dataLabels.showCatName)
        self.assertTrue(chart.dataLabels.showPercent)
        self.assertIsNone(chart.legend)
        ref = chart.series[0].val.numRef.f
        self.assertEqual(ref, f"'{DISTRIBUTION_SHEET}'!$B$2:$B${1 + len(TAXONOMY)}")

    # -- plan: test_mapping_helper_static_content --------------------------
    def test_mapping_helper_static_content(self):
        ws = self.wb[MAPPING_FORM_SHEET]
        got = {
            (c.row, c.column): c.value
            for row in ws.iter_rows()
            for c in row
            if c.value is not None
        }
        expected = {(r, c): v for r, c, v in MAPPING_FORM_CELLS}
        self.assertEqual(got, expected)
        self.assertEqual(ws.cell(1, 1).value, "כמות קופת חולים ")
        self.assertEqual(ws.cell(60, 1).value, 'גמ"ח ____________________')

    # -- plan: test_goals_blank_rows_stay_blank ----------------------------
    def test_goals_blank_rows_stay_blank(self):
        ws = self.wb[GOALS_SHEET]
        self.assertEqual(ws.cell(1, 3).value, "סכום היעד ")
        self.assertEqual(ws.cell(1, 5).value, "גיל היעד ")
        for r in range(3, 9):
            self.assertEqual(ws.cell(r, 6).value, f'=IF(D{r}="","",$F$1-D{r})')
            self.assertEqual(ws.cell(r, 8).value, f'=IF(D{r}="","",F{r}*12)')
            self.assertEqual(ws.cell(r, 10).value, f'=IF(D{r}="","",IF(H{r}>0,$D$1/H{r},""))')
        self.assertEqual(ws.cell(10, 10).value, "=SUM(J3:J8)")
        # no unguarded division anywhere on the sheet
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and "/H" in c.value:
                    self.assertTrue(c.value.startswith('=IF(D'), c.value)

    # -- plan: test_size_under_caps ----------------------------------------
    def test_size_under_caps(self):
        txns = []
        for i in range(5000):
            m = f"2026-{(i % 4) + 1:02d}"
            txns.append(txn(m, (i % 28) + 1, f"עסק {i}", 1000 + i, "מזון ומכולת"))
        wb, path = build(TAXONOMY, txns)
        try:
            with open(path, "rb") as f:
                raw = f.read()
            self.assertLess(len(raw), 15_000_000)
            self.assertLess(len(base64.b64encode(raw)), 20_000_000)
        finally:
            os.unlink(path)


if __name__ == "__main__":
    unittest.main()
